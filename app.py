# app.py
import io
import re
import traceback
import streamlit as st
import pandas as pd
from consultaBD import RepositorioPrincipal
import os
from datetime import datetime
import tempfile
import logging
import time
from settings import FOLDER_LOG_PATH_LOCAL
from settings import FOLDER_LOG_PATH
from settings import SMTP_SERVER
from settings import SMTP_PORT
from settings import SMTP_USER
import openpyxl
import smtplib
from email.message import EmailMessage

# Configuração do sistema de logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# ---> Funções de Log CSV
def encontra_ultimo_arquivo(folder_path, base_name):
    try:
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            logging.info(f"Diretório de log CSV criado: {folder_path}")
        arquivos = os.listdir(folder_path)
        arquivos_data = [
            file
            for file in arquivos
            if re.match(f"^{re.escape(base_name)}(\\d+)\\.csv$", file, re.IGNORECASE)
        ]
        if not arquivos_data:
            logging.info(f"Nenhum log CSV anterior encontrado para base: {base_name}")
            return None
        return max(
            arquivos_data,
            key=lambda x: int(
                re.match(
                    f"^{re.escape(base_name)}(\\d+)\\.csv$", x, re.IGNORECASE
                ).group(1)
            ),
        )
    except Exception as e:
        logging.error(f"Erro ao encontrar último arquivo de log CSV: {e}")
        return None

def cria_proximo_arquivo(folder_path, base_name):
    try:
        ultimo_arquivo = encontra_ultimo_arquivo(folder_path, base_name)
        ultimo_numero = (
            int(
                re.match(
                    f"^{re.escape(base_name)}(\\d+)\\.csv$",
                    ultimo_arquivo,
                    re.IGNORECASE,
                ).group(1)
            )
            if ultimo_arquivo
            else 0
        )
        proximo_numero = ultimo_numero + 1
        return os.path.join(folder_path, f"{base_name}{proximo_numero}.csv")
    except Exception as e:
        logging.error(f"Erro ao criar próximo nome de arquivo de log CSV: {e}")
        return os.path.join(folder_path, f"{base_name}1.csv")  # Fallback

def criar_log_csv(quantidade_itens: int, usuario: str, folder_path: str):
    base_name = "data"
    try:
        total_tempo_humano_por_item = 180
        tempo_bot_fixo_segundos = 20

        df_log = pd.DataFrame(
            {
                "Usuario": [usuario],
                "Rotina": "Vendas - After Market",
                "Data/Hora": [time.strftime("%Y-%m-%d %H:%M:%S")],
                "Quantidade de itens": [quantidade_itens],
                "Tempo_humano(segundos)": [
                    quantidade_itens * total_tempo_humano_por_item
                ],
                "Tempo_bot(segundos)": [tempo_bot_fixo_segundos],
            }
        )

        caminho_do_log = cria_proximo_arquivo(
            folder_path=folder_path, base_name=base_name
        )

        if not caminho_do_log:
            logging.error("Erro: Não foi possível gerar um nome de arquivo de log CSV.")
            return

        df_log.to_csv(caminho_do_log, index=False, sep=";", encoding="utf-8-sig")
        logging.info(f"Log CSV de uso salvo com sucesso em: {caminho_do_log}")

    except Exception as e:
        logging.error(f"Erro ao salvar log CSV: {e}")
        traceback.print_exc()

# --- Função helper para verificar bloqueio ---
def verificar_bloqueio_arquivo(
    caminho_arquivo: str, tentativas: int = 5, espera: int = 2
) -> bool:
    if not os.path.exists(caminho_arquivo):
        logging.info(
            f"Arquivo de log Excel não existe em {caminho_arquivo}. Será criado."
        )
        return False

    tentativa_atual = 0
    while tentativa_atual < tentativas:
        try:
            os.rename(caminho_arquivo, caminho_arquivo)
            logging.info(
                f"Verificação de bloqueio: Arquivo {caminho_arquivo} está livre."
            )
            return False
        except (IOError, OSError, PermissionError) as e:
            if (
                hasattr(e, "errno")
                and (e.errno == 13 or e.errno == 32)
                or isinstance(e, PermissionError)
            ):
                logging.warning(
                    f"Verificação de bloqueio (tentativa {tentativa_atual + 1}/{tentativas}): "
                    f"Arquivo {caminho_arquivo} está bloqueado. Tentando novamente em {espera}s..."
                )
                time.sleep(espera)
                tentativa_atual += 1
            else:
                logging.error(
                    f"Erro inesperado ao verificar bloqueio de {caminho_arquivo}: {e}"
                )
                raise

    logging.error(
        f"ARQUIVO BLOQUEADO: {caminho_arquivo} continua bloqueado após {tentativas} tentativas."
    )
    st.toast(
        f"Não foi possível acessar a planilha '{os.path.basename(caminho_arquivo)}'. Verifique se alguém a está usando.",
        icon="❌",
    )
    return True

# --- Função para enviar as linhas selecionadas para o Excel ---
def enviar_para_excel(df_selecionado: pd.DataFrame, loja_filtrada: str) -> None:
    caminho_arquivo = os.path.normpath(FOLDER_LOG_PATH_LOCAL)
    nome_aba = "Base - AfterMarket"

    if df_selecionado.empty:
        logging.info("Nenhuma linha selecionada para adicionar ao log Excel.")
        return

    coluna_nf = f"Última NF {loja_filtrada}"
    coluna_preco = f"Preço Venda {loja_filtrada}"

    df_para_log_temp = df_selecionado.copy()
    colunas_data_log = [
        col for col in df_para_log_temp.columns if "NF" in col or "Previsão" in col
    ]
    for col in colunas_data_log:
        if pd.api.types.is_datetime64_any_dtype(df_para_log_temp[col]):
            df_para_log_temp[col] = df_para_log_temp[col].dt.strftime("%d/%m/%Y")
        elif pd.api.types.is_object_dtype(df_para_log_temp[col]):
            df_para_log_temp[col] = pd.to_datetime(
                df_para_log_temp[col], errors="coerce"
            ).dt.strftime("%d/%m/%Y")

    df_para_log = pd.DataFrame()
    try:
        df_para_log["PN Voss"] = df_para_log_temp["PN Voss"].astype(str)
        df_para_log["PN Cliente"] = df_para_log_temp["PN Cliente"].astype(str)
        df_para_log["Planta"] = df_para_log_temp["Planta"]
        df_para_log["Loja"] = loja_filtrada
        df_para_log["Ultima NF"] = df_para_log_temp[coluna_nf]
        df_para_log["Preço atual"] = df_para_log_temp[coluna_preco]
        df_para_log["Data"] = datetime.now().strftime("%d/%m/%Y")
    except KeyError as e:
        msg_erro = f"Erro de mapeamento no log. A coluna {e} (baseada na loja '{loja_filtrada}') não foi encontrada."
        logging.error(msg_erro)
        st.error(f"{msg_erro} Verifique os nomes das colunas.")
        return
    except Exception as e:
        st.error(f"Erro inesperado ao preparar dados para o log: {e}")
        return

    caminho_diretorio = os.path.dirname(caminho_arquivo)
    if not os.path.exists(caminho_diretorio):
        try:
            os.makedirs(caminho_diretorio, exist_ok=True)
            logging.info(f"Diretório {caminho_diretorio} verificado/criado.")
        except Exception as e_mkdir:
            st.error(
                f"Não foi possível criar o diretório de log: {caminho_diretorio}. Erro: {e_mkdir}"
            )
            return

    logging.info(f"Iniciando verificação de bloqueio para: {caminho_arquivo}")

    if verificar_bloqueio_arquivo(caminho_arquivo):
        return

    try:
        if os.path.exists(caminho_arquivo):
            workbook = openpyxl.load_workbook(caminho_arquivo)
        else:
            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        if nome_aba in workbook.sheetnames:
            sheet = workbook[nome_aba]
        else:
            sheet = workbook.create_sheet(title=nome_aba)
            headers = list(df_para_log.columns)
            sheet.append(headers)

        for _, row in df_para_log.iterrows():
            row_data = list(row)
            try:
                pn_voss_idx = list(df_para_log.columns).index("PN Voss")
                row_data[pn_voss_idx] = (
                    str(row_data[pn_voss_idx])
                    if row_data[pn_voss_idx] is not None
                    else ""
                )
            except ValueError:
                pass

            try:
                pn_cli_idx = list(df_para_log.columns).index("PN Cliente")
                row_data[pn_cli_idx] = (
                    str(row_data[pn_cli_idx])
                    if row_data[pn_cli_idx] is not None
                    else ""
                )
            except ValueError:
                pass

            sheet.append(row_data)

        workbook.save(caminho_arquivo)
        st.success(
            f"{len(df_para_log)} linha(s) adicionada(s) ao log '{nome_aba}' com sucesso!"
        )

    except Exception as e:
        st.error(f"Erro ao salvar no Excel (openpyxl): {e}")
        logging.error(f"Erro ao salvar log Excel (openpyxl): {e}")

# --- Função para identificar o usuário
def obter_nome_usuario() -> str:
    try:
        # 1. Tenta pegar o usuário da URL (passado pelo IIS/bootstrap.aspx)
        usuario_bruto = st.query_params.get("user") # Ajustado para 'user' conforme seu bootstrap.aspx
        if not usuario_bruto:
             usuario_bruto = st.query_params.get("ad_user") # Mantém compatibilidade retroativa

        if isinstance(usuario_bruto, list):
            usuario_bruto = usuario_bruto[0]

        if usuario_bruto:
            # Remove domínio genérico se presente
            usuario_limpo = re.sub(r'.*\\', '', usuario_bruto)
            logging.info(f"Usuário identificado via URL: {usuario_limpo}")
            return usuario_limpo

        # 2. Fallback para variáveis de ambiente locais
        logging.warning("Parâmetro de usuário não encontrado na URL. Tentando fallback local.")
        usuario_local = os.environ.get("USERNAME") or os.environ.get("USER")

        if usuario_local:
            logging.info(f"Usuário identificado via variável de ambiente: {usuario_local}")
            return usuario_local

    except Exception as e:
        logging.error(f"Erro inesperado ao obter nome de usuário: {e}")

    return "Usuário desconhecido"

# --- Função enviar email ---
def enviar_email_notificacao(
    destinatario_principal: str, df_principal: pd.DataFrame, df_selecao: pd.DataFrame
):
    nome_do_usuario_logado = st.session_state.get("nome_usuario")

    if not nome_do_usuario_logado or nome_do_usuario_logado == "Usuário desconhecido":
        st.error("Não foi possível enviar o email, usuário da sessão não identificado.")
        return

    if df_principal.empty and df_selecao.empty:
        st.warning("Não há dados para enviar no e-mail.")
        return

    servidor_smtp = SMTP_SERVER
    porta_smtp_str = SMTP_PORT
    usuario_email = SMTP_USER

    if not all([servidor_smtp, porta_smtp_str, usuario_email]):
        st.error("Configurações de SMTP incompletas. Verifique o .env")
        return

    try:
        porta_smtp_int = int(porta_smtp_str)
    except (ValueError, TypeError):
        st.error(f"Porta SMTP inválida: {porta_smtp_str}")
        return

    temp_file_paths = []
    try:
        if not df_principal.empty:
             with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="After_Market_Filtrado_") as tmp:
                temp_file_paths.append(tmp.name)
                df_principal.to_excel(tmp.name, index=False) # Simplificado para o exemplo

        if not df_selecao.empty:
             with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="After_Market_Selecao_") as tmp:
                temp_file_paths.append(tmp.name)
                df_selecao.to_excel(tmp.name, index=False) # Simplificado para o exemplo

        # --- PREPARAÇÃO DO E-MAIL ---
        msg = EmailMessage()
        msg["Subject"] = f"Análise de After Market - {datetime.now().strftime('%d/%m/%Y')}"
        msg["From"] = usuario_email
        msg["To"] = destinatario_principal
        msg["Cc"] = f"{nome_do_usuario_logado}@example.com"

        msg.set_content("Segue em anexo o relatório de análise de After Market.\n\nAtenciosamente,\nBot")

        for path in temp_file_paths:
            with open(path, 'rb') as f:
                msg.add_attachment(f.read(), maintype='application', subtype='xlsx', filename=os.path.basename(path))

        # --- ENVIO ---
        with smtplib.SMTP(servidor_smtp, porta_smtp_int) as smtp:
            smtp.send_message(msg)

        st.success(f"E-mail enviado para {destinatario_principal}!")

    except Exception as e:
        st.error(f"Erro ao enviar e-mail: {e}")
        logging.error(f"Erro email: {e}")
    finally:
        for path in temp_file_paths:
            if os.path.exists(path):
                os.remove(path)